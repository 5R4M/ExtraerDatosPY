from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener el nombre de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def find_column_titles(ws, search_rows=20):
    titles = {}
    for col in ws.iter_cols(1, ws.max_column):
        for row in range(1, search_rows + 1):
            cell = col[row - 1]
            if cell.value and isinstance(cell, MergedCell):
                cell = ws.cell(row=cell.row, column=cell.column)
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value not in titles:  # Evitar duplicados
                    titles[cell_value] = get_column_letter(cell.column)
    return titles

def extract_and_transfer(ws_origen, columnas_seleccionadas):
    try:
        if not columnas_seleccionadas:
            messagebox.showerror("Error", "Debe seleccionar al menos una columna para extraer datos.")
            return
        
        # Crear un nuevo archivo de Excel
        wb = Workbook()
        ws = wb.active

        # Escribir los títulos de las columnas seleccionadas en el nuevo archivo
        for c_idx, col_title in enumerate(list(columnas_seleccionadas) + ['Etiqueta'], 1):
            cell = ws.cell(row=1, column=c_idx, value=col_title.title())
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border
            cell.font = Font(name='Arial', size=12, bold=True)

        # Extraer datos específicos según las columnas seleccionadas
        datos_extraidos = []
        for row in ws_origen.iter_rows(min_row=2, max_row=ws_origen.max_row):
            fila = {}
            for col_title, col_letter in columnas_seleccionadas.items():
                cell_value = row[ws_origen[col_letter + '1'].column - 1].value
                fila[col_title] = cell_value
            if any(fila.values()):
                datos_extraidos.append(fila)

        # Crear una nueva columna combinada con los títulos
        for fila in datos_extraidos:
            fila['Etiqueta'] = ' '.join([f"{titulo.title()}: {fila[titulo]}" for titulo in columnas_seleccionadas])

        # Escribir los datos extraídos en el nuevo archivo
        for r_idx, fila in enumerate(datos_extraidos, 2):
            for c_idx, (col, value) in enumerate(fila.items(), 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.border = thin_border
                cell.font = Font(name='Arial', size=12)

        # Ajustar el tamaño de las celdas al contenido
        adjust_column_width(ws)

        # Guardar el nuevo archivo de Excel
        destino = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if destino:
            wb.save(destino)
            messagebox.showinfo("Éxito", "Datos transferidos exitosamente")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def seleccionar_hoja(wb_origen, hojas, root):
    def on_selection():
        hoja_seleccionada = combobox.get()
        if hoja_seleccionada:
            seleccionar_columnas(wb_origen[hoja_seleccionada])
            ventana.destroy()

    ventana = tk.Toplevel(root)
    ventana.title("Seleccionar Hoja")

    label = tk.Label(ventana, text="Seleccione la hoja de trabajo:")
    label.pack(pady=10)

    combobox = ttk.Combobox(ventana, values=hojas, state="readonly")
    combobox.pack(pady=10)

    boton = tk.Button(ventana, text="Aceptar", command=on_selection)
    boton.pack(pady=20)

def seleccionar_columnas(ws_origen):
    def on_columns_selected():
        columnas_seleccionadas = listbox.curselection()
        selected_columns = {listbox.get(i): column_titles[listbox.get(i)] for i in columnas_seleccionadas}
        if selected_columns:
            extract_and_transfer(ws_origen, selected_columns)
        ventana.destroy()

    ventana = tk.Toplevel(root)
    ventana.title("Seleccionar Columnas")

    label = tk.Label(ventana, text="Seleccione las columnas para extraer datos:")
    label.pack(pady=10)

    # Obtener los títulos de las columnas de las primeras 20 filas
    column_titles = find_column_titles(ws_origen)

    listbox = tk.Listbox(ventana, selectmode="multiple")
    for title in column_titles:
        listbox.insert(tk.END, title)
    listbox.pack(pady=10)

    boton = tk.Button(ventana, text="Aceptar", command=on_columns_selected)
    boton.pack(pady=20)

def abrir_archivo():
    origen = filedialog.askopenfilename(title="Seleccionar archivo de origen", filetypes=[("Excel files", "*.xlsx")])
    if origen:
        wb_origen = load_workbook(origen)
        hojas = wb_origen.sheetnames
        seleccionar_hoja(wb_origen, hojas, root)

# Crear la ventana principal
root = tk.Tk()
root.title("Transferencia de Datos de Excel")

# Crear el botón para iniciar la selección de archivo
btn_abrir = tk.Button(root, text="Abrir Archivo y Seleccionar Hoja", command=abrir_archivo)
btn_abrir.pack(pady=20)

# Ejecutar la aplicación
root.mainloop()